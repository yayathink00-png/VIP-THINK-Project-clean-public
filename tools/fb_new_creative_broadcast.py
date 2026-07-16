#!/usr/bin/env python3
"""
fb_new_creative_broadcast.py

Read-only Meta Ads API monitor for newly seen Facebook ad creatives.

It reads ads/insights from Meta, stores a local SQLite history, and writes a
Markdown + CSV + optional Excel broadcast for creatives first seen by this
monitor. It does not create, edit, pause, or publish ads.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover - fallback for minimal environments
    load_dotenv = None

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover - CSV/Markdown still work
    Workbook = None


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_API_VERSION = "v21.0"
DEFAULT_DB_NAME = "fb_new_creative_monitor.sqlite3"
DEFAULT_OUTPUT_DIR = "outputs/fb_new_creative_broadcast"
TRANSIENT_CODES = {1, 2, 4, 17, 32, 613}
DEFAULT_LEAD_ACTION_TYPES = {
    "lead",
    "onsite_conversion.lead_grouped",
    "offsite_conversion.fb_pixel_lead",
    "offsite_conversion.fb_pixel_complete_registration",
    "complete_registration",
    "schedule",
    "submit_application",
}

AD_FIELDS_FULL = (
    "id,name,created_time,updated_time,status,effective_status,configured_status,"
    "creative{id,name,thumbnail_url,object_type,object_story_spec,asset_feed_spec,"
    "image_hash,video_id,body,title},"
    "adset{id,name,start_time,end_time,status,effective_status},"
    "campaign{id,name,start_time,stop_time,status,effective_status}"
)
AD_FIELDS_LEAN = (
    "id,name,created_time,updated_time,status,effective_status,configured_status,"
    "creative{id,name,thumbnail_url,object_story_spec,asset_feed_spec},"
    "adset{id,name,start_time,end_time,status,effective_status},"
    "campaign{id,name,start_time,stop_time,status,effective_status}"
)
AD_SCAN_FIELDS = (
    "id,name,created_time,updated_time,status,effective_status,configured_status,"
    "adset{id,name,start_time,end_time,status,effective_status},"
    "campaign{id,name,start_time,stop_time,status,effective_status}"
)
INSIGHT_FIELDS = (
    "ad_id,ad_name,adset_id,adset_name,campaign_id,campaign_name,"
    "spend,impressions,clicks,actions,date_start,date_stop"
)


@dataclass(frozen=True)
class Config:
    access_token: str
    accounts: list[str]
    api_version: str


class GraphAPIError(RuntimeError):
    pass


class FBClient:
    def __init__(self, token: str, api_version: str) -> None:
        self.token = token
        self.api_version = api_version
        self.base_url = f"https://graph.facebook.com/{api_version}"

    def get(self, path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        return self._request("GET", path, params=params)

    def paged(self, path: str, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        next_path: str | None = path
        first = True
        while next_path:
            if first:
                data = self.get(next_path, params=params)
                first = False
            else:
                data = self._request_url(next_path)
            rows.extend(data.get("data", []))
            next_path = data.get("paging", {}).get("next")
        return rows

    def _request(self, method: str, path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        url = path if path.startswith("http") else f"{self.base_url}/{path.lstrip('/')}"
        query = dict(params or {})
        query.setdefault("access_token", self.token)
        return self._send(method, url, params=query)

    def _request_url(self, url: str) -> dict[str, Any]:
        return self._send("GET", url, params=None)

    def _send(self, method: str, url: str, params: dict[str, Any] | None) -> dict[str, Any]:
        last_error = ""
        for attempt in range(5):
            request_url = append_query(url, params) if params else url
            request = urllib.request.Request(
                request_url,
                headers={"User-Agent": "fb-new-creative-broadcast/0.1"},
                method=method,
            )
            try:
                with urllib.request.urlopen(request, timeout=90) as resp:
                    body = resp.read().decode("utf-8")
                    return json.loads(body)
            except urllib.error.HTTPError as exc:
                body = exc.read().decode("utf-8", errors="replace")
                err = _extract_error(body)
                code = err.get("code")
                if exc.code in {429, 500, 502, 503, 504} or code in TRANSIENT_CODES:
                    if attempt < 4:
                        time.sleep((2**attempt) * 2)
                        continue
                safe_message = _redact_secret(json.dumps(err, ensure_ascii=False), self.token)
                raise GraphAPIError(f"HTTP {exc.code}: {safe_message}") from exc
            except urllib.error.URLError as exc:
                last_error = str(exc.reason)
                if attempt < 4:
                    time.sleep(2**attempt)
                    continue
                raise GraphAPIError(f"Graph API connection failed: {last_error}") from exc
            except json.JSONDecodeError as exc:
                raise GraphAPIError("Graph API returned non-JSON response") from exc

        raise GraphAPIError(f"Graph API failed after retries: {_redact_secret(last_error, self.token)}")


def append_query(url: str, params: dict[str, Any]) -> str:
    encoded = urllib.parse.urlencode({k: encode_param_value(v) for k, v in params.items()})
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}{encoded}"


def encode_param_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _extract_error(body: str) -> dict[str, Any]:
    try:
        payload = json.loads(body)
        err = payload.get("error")
        if isinstance(err, dict):
            return err
        return payload if isinstance(payload, dict) else {"message": str(payload)}
    except Exception:
        return {"message": body[:500]}


def _redact_secret(text: str, token: str) -> str:
    if token:
        text = text.replace(token, "REDACTED")
    return re.sub(r"access_token=[^&\s]+", "access_token=REDACTED", text)


def load_environment() -> None:
    if load_dotenv:
        load_dotenv(SCRIPT_DIR / ".env")
        load_dotenv(Path.cwd() / ".env")
        return
    env_path = SCRIPT_DIR / ".env"
    if not env_path.exists():
        return
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def parse_accounts(raw: str) -> list[str]:
    accounts: list[str] = []
    for item in re.split(r"[,;\s]+", raw.strip()):
        account = normalize_account_id(item)
        if account and account not in accounts:
            accounts.append(account)
    return accounts


def normalize_account_id(value: str) -> str:
    value = str(value or "").strip()
    if not value:
        return ""
    if value.startswith("act_"):
        return value
    if value.isdigit():
        return f"act_{value}"
    return value


def load_config(args: argparse.Namespace) -> Config:
    load_environment()
    token = os.getenv("FB_ACCESS_TOKEN", "").strip()
    if not token:
        sys.exit("[ERR] Missing FB_ACCESS_TOKEN in .env or environment.")

    account_source = (
        args.accounts
        or os.getenv("FB_ALL_AD_ACCOUNTS", "")
        or os.getenv("FB_AD_ACCOUNT_IDS", "")
        or os.getenv("FB_AD_ACCOUNT_ID", "")
    )
    accounts = parse_accounts(account_source)
    if not accounts:
        sys.exit("[ERR] Missing ad account IDs. Set FB_ALL_AD_ACCOUNTS or FB_AD_ACCOUNT_ID.")

    api_version = (args.api_version or os.getenv("FB_API_VERSION") or DEFAULT_API_VERSION).strip()
    return Config(access_token=token, accounts=accounts, api_version=api_version)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a read-only FB new creative broadcast from Meta Ads API."
    )
    parser.add_argument("--days", type=int, default=14, help="Lookback window ending at --until. Default: 14.")
    parser.add_argument("--since", help="Start date YYYY-MM-DD. Overrides --days.")
    parser.add_argument("--until", help="End date YYYY-MM-DD. Default: today.")
    parser.add_argument("--accounts", help="Comma/space separated account IDs. Defaults to FB_ALL_AD_ACCOUNTS.")
    parser.add_argument("--api-version", help="Meta API version. Defaults to FB_API_VERSION or v21.0.")
    parser.add_argument("--db", default=DEFAULT_DB_NAME, help="SQLite history path.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Report output directory.")
    parser.add_argument(
        "--source-mode",
        choices=["created-or-scheduled", "activity"],
        default="created-or-scheduled",
        help=(
            "created-or-scheduled reports ads created in the window or scheduled via adset start_time. "
            "activity uses ads with Insights activity. Default: created-or-scheduled."
        ),
    )
    parser.add_argument("--baseline-only", action="store_true", help="Seed history but do not report new creatives.")
    parser.add_argument("--no-save", action="store_true", help="Do not write SQLite history; useful for testing.")
    parser.add_argument(
        "--skip-zero-spend-scan",
        action="store_true",
        help="Only use Insights rows. Faster, but misses newly created ads with zero spend/impressions.",
    )
    parser.add_argument(
        "--max-ads-per-account",
        type=int,
        default=5000,
        help="Safety cap for zero-spend ad scan. 0 means unlimited. Default: 5000.",
    )
    parser.add_argument(
        "--future-scheduled-days",
        type=int,
        default=14,
        help="Also include ads whose adset start_time is in the next N days. Default: 14.",
    )
    parser.add_argument(
        "--skip-insights",
        action="store_true",
        help="In created-or-scheduled mode, skip performance metrics and only report ad/creative metadata.",
    )
    parser.add_argument(
        "--min-spend-alert",
        type=float,
        default=300.0,
        help="Flag new creatives with spend above this and no leads. Default: 300.",
    )
    parser.add_argument(
        "--lead-action-types",
        default=os.getenv("FB_LEAD_ACTION_TYPES", ""),
        help="Comma-separated action_type values counted as leads/appointments.",
    )
    parser.add_argument(
        "--daily-insights",
        action="store_true",
        help="Fetch Insights with time_increment=1. Slower; default uses one aggregate row per ad.",
    )
    parser.add_argument("--check-only", action="store_true", help="Only verify token/account read access.")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def resolve_window(args: argparse.Namespace) -> tuple[date, date]:
    until = parse_date(args.until) if args.until else date.today()
    if args.since:
        since = parse_date(args.since)
    else:
        days = max(args.days, 1)
        since = until - timedelta(days=days - 1)
    if since > until:
        sys.exit("[ERR] --since cannot be later than --until.")
    return since, until


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        sys.exit(f"[ERR] Invalid date: {value}. Use YYYY-MM-DD.")


def parse_meta_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None
    raw = raw.replace("+0000", "+00:00").replace("-0000", "+00:00")
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        try:
            return datetime.strptime(value[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            return None


def is_between_meta_date(value: str | None, since: date, until: date) -> bool:
    parsed = parse_meta_datetime(value)
    if parsed is None:
        return False
    current = parsed.date()
    return since <= current <= until


def connect_db(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    ensure_schema(conn)
    return conn


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS creatives (
            creative_key TEXT PRIMARY KEY,
            creative_id TEXT,
            creative_name TEXT,
            material_type TEXT,
            material_label TEXT,
            thumbnail_url TEXT,
            image_hash TEXT,
            video_id TEXT,
            object_type TEXT,
            first_seen_at TEXT NOT NULL,
            first_seen_date TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            first_account_id TEXT,
            first_ad_id TEXT,
            first_ad_name TEXT,
            first_campaign_name TEXT,
            first_adset_name TEXT,
            meta_created_time TEXT,
            raw_json TEXT
        );

        CREATE TABLE IF NOT EXISTS ad_daily_metrics (
            account_id TEXT NOT NULL,
            ad_id TEXT NOT NULL,
            creative_key TEXT NOT NULL,
            date_start TEXT NOT NULL,
            date_stop TEXT NOT NULL,
            ad_name TEXT,
            adset_id TEXT,
            adset_name TEXT,
            campaign_id TEXT,
            campaign_name TEXT,
            spend REAL DEFAULT 0,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            leads REAL DEFAULT 0,
            actions_json TEXT,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (account_id, ad_id, creative_key, date_start, date_stop)
        );

        CREATE TABLE IF NOT EXISTS ads_seen (
            ad_id TEXT PRIMARY KEY,
            account_id TEXT NOT NULL,
            creative_key TEXT,
            first_seen_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            ad_name TEXT,
            campaign_name TEXT,
            adset_name TEXT,
            created_time TEXT,
            updated_time TEXT,
            adset_start_time TEXT,
            campaign_start_time TEXT,
            schedule_reason TEXT,
            effective_status TEXT,
            raw_json TEXT
        );

        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            started_at TEXT NOT NULL,
            since_date TEXT NOT NULL,
            until_date TEXT NOT NULL,
            accounts_json TEXT NOT NULL,
            baseline_only INTEGER NOT NULL,
            discovered_creatives INTEGER NOT NULL,
            reported_new_creatives INTEGER NOT NULL,
            report_dir TEXT
        );
        """
    )
    conn.commit()


def existing_creative_keys(conn: sqlite3.Connection | None) -> set[str]:
    if conn is None:
        return set()
    return {str(row["creative_key"]) for row in conn.execute("SELECT creative_key FROM creatives")}


def existing_ad_ids(conn: sqlite3.Connection | None) -> set[str]:
    if conn is None:
        return set()
    return {str(row["ad_id"]) for row in conn.execute("SELECT ad_id FROM ads_seen")}


def fetch_account_name(client: FBClient, account_id: str) -> str:
    data = client.get(account_id, {"fields": "id,name"})
    return str(data.get("name") or account_id)


def fetch_insights(
    client: FBClient,
    account_id: str,
    since: date,
    until: date,
    daily: bool = False,
    verbose: bool = False,
) -> list[dict[str, Any]]:
    params = {
        "fields": INSIGHT_FIELDS,
        "level": "ad",
        "time_range": json.dumps({"since": since.isoformat(), "until": until.isoformat()}),
        "limit": 200,
    }
    if daily:
        params["time_increment"] = 1
    if verbose:
        print(f"[INFO] {account_id}: fetching insights {since} to {until}")
    return client.paged(f"{account_id}/insights", params)


def fetch_ad(client: FBClient, ad_id: str) -> dict[str, Any]:
    try:
        return client.get(ad_id, {"fields": AD_FIELDS_FULL})
    except GraphAPIError:
        return client.get(ad_id, {"fields": AD_FIELDS_LEAN})


def fetch_ads_by_ids(client: FBClient, ad_ids: list[str]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    ads: dict[str, dict[str, Any]] = {}
    errors: list[str] = []
    for chunk in chunks(ad_ids, 50):
        try:
            payload = client.get("", {"ids": ",".join(chunk), "fields": AD_FIELDS_FULL})
        except GraphAPIError:
            try:
                payload = client.get("", {"ids": ",".join(chunk), "fields": AD_FIELDS_LEAN})
            except GraphAPIError as exc:
                errors.append(f"batch {chunk[0]}..{chunk[-1]}: {exc}")
                for ad_id in chunk:
                    try:
                        ads[ad_id] = fetch_ad(client, ad_id)
                    except GraphAPIError as item_exc:
                        errors.append(f"ad {ad_id}: {item_exc}")
                continue
        for ad_id in chunk:
            item = payload.get(ad_id)
            if isinstance(item, dict) and "error" not in item:
                ads[ad_id] = item
            elif isinstance(item, dict):
                errors.append(f"ad {ad_id}: {json.dumps(item.get('error'), ensure_ascii=False)}")
    return ads, errors


def chunks(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def fetch_recent_ads(
    client: FBClient,
    account_id: str,
    since: date,
    until: date,
    max_ads: int,
    verbose: bool = False,
) -> tuple[list[dict[str, Any]], bool]:
    if verbose:
        print(f"[INFO] {account_id}: scanning ads for zero-spend recent creations")
    try:
        ads = client.paged(f"{account_id}/ads", {"fields": AD_FIELDS_FULL, "limit": 200})
    except GraphAPIError:
        ads = client.paged(f"{account_id}/ads", {"fields": AD_FIELDS_LEAN, "limit": 200})

    truncated = False
    if max_ads > 0 and len(ads) > max_ads:
        ads = ads[:max_ads]
        truncated = True

    recent = [
        ad
        for ad in ads
        if is_between_meta_date(ad.get("created_time"), since, until)
        or is_between_meta_date(ad.get("updated_time"), since, until)
    ]
    return recent, truncated


def fetch_created_or_scheduled_ads(
    client: FBClient,
    account_id: str,
    since: date,
    until: date,
    future_scheduled_days: int,
    max_ads: int,
    verbose: bool = False,
) -> tuple[list[dict[str, Any]], bool, list[str]]:
    if verbose:
        print(f"[INFO] {account_id}: scanning ads by created_time/adset.start_time", flush=True)
    errors: list[str] = []
    scan_by_id: dict[str, dict[str, Any]] = {}

    try:
        for ad in fetch_ads_created_in_window(client, account_id, since, until):
            ad_id = str(ad.get("id") or "")
            if ad_id:
                scan_by_id[ad_id] = ad
    except GraphAPIError as exc:
        errors.append(f"created_time filter failed: {exc}")

    try:
        scheduled_ads = fetch_ads_from_scheduled_adsets(
            client,
            account_id,
            since,
            until + timedelta(days=max(future_scheduled_days, 0)),
        )
        for ad in scheduled_ads:
            ad_id = str(ad.get("id") or "")
            if ad_id:
                scan_by_id[ad_id] = merge_scan_rows(scan_by_id.get(ad_id), ad)
    except GraphAPIError as exc:
        errors.append(f"adset start_time filter failed: {exc}")

    ads = list(scan_by_id.values())

    truncated = False
    if max_ads > 0 and len(ads) > max_ads:
        ads = ads[:max_ads]
        truncated = True

    candidate_scan_rows: list[dict[str, Any]] = []
    for ad in ads:
        reason = classify_created_or_scheduled_ad(ad, since, until, future_scheduled_days)
        if not reason:
            continue
        ad["_schedule_reason"] = reason
        candidate_scan_rows.append(ad)

    candidate_ids = [str(ad.get("id")) for ad in candidate_scan_rows if ad.get("id")]
    if not candidate_ids:
        return [], truncated, errors

    detailed_ads, detail_errors = fetch_ads_by_ids(client, candidate_ids)
    errors.extend(detail_errors)

    candidates: list[dict[str, Any]] = []
    scan_by_id = {str(ad.get("id")): ad for ad in candidate_scan_rows if ad.get("id")}
    for ad_id in candidate_ids:
        detail = detailed_ads.get(ad_id) or scan_by_id[ad_id]
        detail["_schedule_reason"] = scan_by_id[ad_id].get("_schedule_reason", "")
        if "adset" not in detail and scan_by_id[ad_id].get("adset"):
            detail["adset"] = scan_by_id[ad_id]["adset"]
        if "campaign" not in detail and scan_by_id[ad_id].get("campaign"):
            detail["campaign"] = scan_by_id[ad_id]["campaign"]
        candidates.append(detail)
    return candidates, truncated, errors


def fetch_ads_created_in_window(
    client: FBClient,
    account_id: str,
    since: date,
    until: date,
) -> list[dict[str, Any]]:
    filtering = [
        {"field": "created_time", "operator": "GREATER_THAN", "value": start_ts(since)},
        {"field": "created_time", "operator": "LESS_THAN", "value": end_ts(until)},
    ]
    return client.paged(
        f"{account_id}/ads",
        {
            "fields": AD_SCAN_FIELDS,
            "filtering": json.dumps(filtering),
            "limit": 200,
        },
    )


def fetch_ads_from_scheduled_adsets(
    client: FBClient,
    account_id: str,
    since: date,
    until: date,
) -> list[dict[str, Any]]:
    filtering = [
        {"field": "start_time", "operator": "GREATER_THAN", "value": start_ts(since)},
        {"field": "start_time", "operator": "LESS_THAN", "value": end_ts(until)},
    ]
    adsets = client.paged(
        f"{account_id}/adsets",
        {
            "fields": "id,name,start_time,end_time,status,effective_status",
            "filtering": json.dumps(filtering),
            "limit": 200,
        },
    )
    ads: list[dict[str, Any]] = []
    for adset in adsets:
        adset_id = str(adset.get("id") or "")
        if not adset_id:
            continue
        for ad in client.paged(f"{adset_id}/ads", {"fields": AD_SCAN_FIELDS, "limit": 200}):
            ad["adset"] = merge_scan_rows(ad.get("adset"), adset)
            ads.append(ad)
    return ads


def start_ts(day: date) -> int:
    return int(datetime.combine(day, datetime.min.time(), tzinfo=timezone.utc).timestamp())


def end_ts(day: date) -> int:
    next_day = day + timedelta(days=1)
    return int(datetime.combine(next_day, datetime.min.time(), tzinfo=timezone.utc).timestamp())


def merge_scan_rows(first: dict[str, Any] | None, second: dict[str, Any] | None) -> dict[str, Any]:
    merged = dict(first or {})
    merged.update({k: v for k, v in dict(second or {}).items() if v not in (None, "")})
    return merged


def classify_created_or_scheduled_ad(
    ad: dict[str, Any],
    since: date,
    until: date,
    future_scheduled_days: int,
) -> str:
    reasons: list[str] = []
    created = parse_meta_datetime(ad.get("created_time"))
    if created and since <= created.date() <= until:
        reasons.append("14天内新建广告")

    adset = ad.get("adset") if isinstance(ad.get("adset"), dict) else {}
    adset_start = parse_meta_datetime(adset.get("start_time"))
    if adset_start:
        start_date = adset_start.date()
        if since <= start_date <= until:
            reasons.append("广告组排期在14天窗口内")
        future_until = until + timedelta(days=max(future_scheduled_days, 0))
        if until < start_date <= future_until:
            reasons.append(f"未来{future_scheduled_days}天已排期")

    return "；".join(reasons)


def action_types(args: argparse.Namespace) -> set[str]:
    raw = args.lead_action_types.strip()
    if not raw:
        return set(DEFAULT_LEAD_ACTION_TYPES)
    return {item.strip() for item in raw.split(",") if item.strip()}


def summarize_insights(rows: list[dict[str, Any]], lead_types: set[str]) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    by_ad: dict[str, dict[str, Any]] = {}
    daily_rows: list[dict[str, Any]] = []
    for row in rows:
        ad_id = str(row.get("ad_id") or "")
        if not ad_id:
            continue
        spend = as_float(row.get("spend"))
        impressions = as_int(row.get("impressions"))
        clicks = as_int(row.get("clicks"))
        leads = extract_leads(row.get("actions"), lead_types)
        entry = by_ad.setdefault(
            ad_id,
            {
                "ad_id": ad_id,
                "ad_name": row.get("ad_name") or "",
                "adset_id": row.get("adset_id") or "",
                "adset_name": row.get("adset_name") or "",
                "campaign_id": row.get("campaign_id") or "",
                "campaign_name": row.get("campaign_name") or "",
                "spend": 0.0,
                "impressions": 0,
                "clicks": 0,
                "leads": 0.0,
            },
        )
        entry["spend"] += spend
        entry["impressions"] += impressions
        entry["clicks"] += clicks
        entry["leads"] += leads
        daily_rows.append(
            {
                "ad_id": ad_id,
                "date_start": row.get("date_start") or "",
                "date_stop": row.get("date_stop") or "",
                "spend": spend,
                "impressions": impressions,
                "clicks": clicks,
                "leads": leads,
                "actions_json": json.dumps(row.get("actions") or [], ensure_ascii=False),
                "ad_name": row.get("ad_name") or "",
                "adset_id": row.get("adset_id") or "",
                "adset_name": row.get("adset_name") or "",
                "campaign_id": row.get("campaign_id") or "",
                "campaign_name": row.get("campaign_name") or "",
            }
        )
    return by_ad, daily_rows


def extract_leads(actions: Any, lead_types: set[str]) -> float:
    if not isinstance(actions, list):
        return 0.0
    total = 0.0
    for action in actions:
        if not isinstance(action, dict):
            continue
        if str(action.get("action_type") or "") in lead_types:
            total += as_float(action.get("value"))
    return total


def as_float(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_int(value: Any) -> int:
    try:
        if value is None or value == "":
            return 0
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def creative_from_ad(account_id: str, account_name: str, ad: dict[str, Any], metrics: dict[str, Any]) -> dict[str, Any] | None:
    creative = ad.get("creative") if isinstance(ad.get("creative"), dict) else {}
    if not creative:
        return None

    media = extract_media_refs(creative)
    creative_id = str(creative.get("id") or "")
    creative_name = str(creative.get("name") or metrics.get("ad_name") or ad.get("name") or "")
    ad_name = str(ad.get("name") or metrics.get("ad_name") or "")
    adset = ad.get("adset") if isinstance(ad.get("adset"), dict) else {}
    campaign = ad.get("campaign") if isinstance(ad.get("campaign"), dict) else {}
    adset_name = str(adset.get("name") or metrics.get("adset_name") or "")
    campaign_name = str(campaign.get("name") or metrics.get("campaign_name") or "")
    adset_start_time = str(adset.get("start_time") or "")
    campaign_start_time = str(campaign.get("start_time") or "")

    creative_key = build_creative_key(account_id, creative_id, media, creative_name or ad_name)
    if not creative_key:
        return None

    material_type = classify_material_type(media, creative, creative_name, ad_name)
    material_label = classify_material_label(material_type, creative_name, ad_name)
    region = infer_region(" ".join([account_name, campaign_name, adset_name, ad_name, creative_name]))
    thumbnail = media.get("thumbnail_url") or str(creative.get("thumbnail_url") or "")
    spend = as_float(metrics.get("spend"))
    impressions = as_int(metrics.get("impressions"))
    clicks = as_int(metrics.get("clicks"))
    leads = as_float(metrics.get("leads"))

    return {
        "creative_key": creative_key,
        "creative_id": creative_id,
        "creative_name": creative_name,
        "material_type": material_type,
        "material_label": material_label,
        "thumbnail_url": thumbnail,
        "image_hash": media.get("image_hash") or "",
        "video_id": media.get("video_id") or "",
        "object_type": str(creative.get("object_type") or ""),
        "account_id": account_id,
        "account_name": account_name,
        "ad_id": str(ad.get("id") or metrics.get("ad_id") or ""),
        "ad_name": ad_name,
        "adset_id": str(adset.get("id") or metrics.get("adset_id") or ""),
        "adset_name": adset_name,
        "campaign_id": str(campaign.get("id") or metrics.get("campaign_id") or ""),
        "campaign_name": campaign_name,
        "created_time": str(ad.get("created_time") or ""),
        "updated_time": str(ad.get("updated_time") or ""),
        "adset_start_time": adset_start_time,
        "campaign_start_time": campaign_start_time,
        "schedule_reason": str(ad.get("_schedule_reason") or ""),
        "status": str(ad.get("status") or ""),
        "effective_status": str(ad.get("effective_status") or ""),
        "region": region,
        "spend": spend,
        "impressions": impressions,
        "clicks": clicks,
        "ctr": (clicks / impressions) if impressions else 0.0,
        "leads": leads,
        "raw_json": json.dumps({"ad": ad, "creative": creative}, ensure_ascii=False),
    }


def extract_media_refs(creative: dict[str, Any]) -> dict[str, str]:
    media: dict[str, str] = {}
    for key in ("image_hash", "video_id", "thumbnail_url"):
        if creative.get(key):
            media[key] = str(creative[key])

    oss = parse_jsonish(creative.get("object_story_spec")) or {}
    link_data = parse_jsonish(oss.get("link_data")) or {}
    video_data = parse_jsonish(oss.get("video_data")) or {}
    if link_data.get("image_hash") and not media.get("image_hash"):
        media["image_hash"] = str(link_data["image_hash"])
    if link_data.get("picture") and not media.get("thumbnail_url"):
        media["thumbnail_url"] = str(link_data["picture"])
    if video_data.get("video_id") and not media.get("video_id"):
        media["video_id"] = str(video_data["video_id"])
    if video_data.get("image_url") and not media.get("thumbnail_url"):
        media["thumbnail_url"] = str(video_data["image_url"])

    afs = parse_jsonish(creative.get("asset_feed_spec")) or {}
    videos = afs.get("videos") if isinstance(afs.get("videos"), list) else []
    images = afs.get("images") if isinstance(afs.get("images"), list) else []
    if videos and isinstance(videos[0], dict):
        if videos[0].get("video_id") and not media.get("video_id"):
            media["video_id"] = str(videos[0]["video_id"])
        if videos[0].get("thumbnail_url") and not media.get("thumbnail_url"):
            media["thumbnail_url"] = str(videos[0]["thumbnail_url"])
    if images and isinstance(images[0], dict):
        if images[0].get("hash") and not media.get("image_hash"):
            media["image_hash"] = str(images[0]["hash"])
        if images[0].get("url") and not media.get("thumbnail_url"):
            media["thumbnail_url"] = str(images[0]["url"])
    return media


def parse_jsonish(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str) and value.strip().startswith(("{", "[")):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return {}
    return {}


def build_creative_key(account_id: str, creative_id: str, media: dict[str, str], name: str) -> str:
    if creative_id:
        return f"creative:{creative_id}"
    if media.get("video_id"):
        return f"video:{media['video_id']}"
    if media.get("image_hash"):
        return f"image:{account_id}:{media['image_hash']}"
    cleaned = normalize_name(name)
    if cleaned:
        return f"name:{account_id}:{cleaned}"
    return ""


def normalize_name(value: str) -> str:
    text = re.sub(r"\s+", " ", value.strip().lower())
    text = re.sub(r"\.(mp4|mov|avi|mkv|webm|m4v|jpg|jpeg|png|gif|webp|bmp)$", "", text)
    return text


def classify_material_type(media: dict[str, str], creative: dict[str, Any], creative_name: str, ad_name: str) -> str:
    joined = f"{creative_name} {ad_name}".lower()
    if media.get("video_id") or "video_data" in json.dumps(creative, ensure_ascii=False).lower():
        return "video"
    if media.get("image_hash") or "link_data" in json.dumps(creative, ensure_ascii=False).lower():
        return "image"
    if any(ext in joined for ext in (".mp4", ".mov", ".m4v", "视频", "video")):
        return "video"
    if any(ext in joined for ext in (".jpg", ".jpeg", ".png", "图片", "图文")):
        return "image"
    return "unknown"


def classify_material_label(material_type: str, creative_name: str, ad_name: str) -> str:
    joined = f"{creative_name} {ad_name}".lower()
    if "ai视频" in joined or "ai video" in joined or "ai_video" in joined:
        return "AI视频"
    if "sl视频" in joined or "sl video" in joined or "sl_video" in joined:
        return "SL视频"
    if material_type == "video":
        return "视频"
    if material_type == "image":
        return "图片"
    return "未识别"


def infer_region(text: str) -> str:
    lowered = text.lower()
    if re.search(r"台湾|台灣|taiwan|\btw\b", lowered):
        return "台湾"
    if re.search(r"香港|港澳|hong\s*kong|\bhk\b|澳门|澳門|macau", lowered):
        return "香港/港澳"
    return "未识别"


def merge_creative_record(target: dict[str, Any], item: dict[str, Any]) -> None:
    target["spend"] += item["spend"]
    target["impressions"] += item["impressions"]
    target["clicks"] += item["clicks"]
    target["leads"] += item["leads"]
    target["ctr"] = (target["clicks"] / target["impressions"]) if target["impressions"] else 0.0
    target["ads_count"] += 1
    target["account_ids"].add(item["account_id"])
    target["account_names"].add(item["account_name"])
    target["ad_names"].add(item["ad_name"])
    target["campaign_names"].add(item["campaign_name"])
    target["adset_names"].add(item["adset_name"])
    target["schedule_reasons"].add(item["schedule_reason"])
    if target["region"] == "未识别" and item["region"] != "未识别":
        target["region"] = item["region"]
    if target["material_label"] == "未识别" and item["material_label"] != "未识别":
        target["material_label"] = item["material_label"]


def build_records(
    account_id: str,
    account_name: str,
    ads: dict[str, dict[str, Any]],
    metrics_by_ad: dict[str, dict[str, Any]],
    include_metric_only: bool = True,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    records: dict[str, dict[str, Any]] = {}
    details: list[dict[str, Any]] = []
    ad_ids = set(ads)
    if include_metric_only:
        ad_ids |= set(metrics_by_ad)
    for ad_id in sorted(ad_ids):
        ad = ads.get(ad_id) or {}
        metrics = metrics_by_ad.get(ad_id) or {"ad_id": ad_id}
        item = creative_from_ad(account_id, account_name, ad, metrics)
        if not item:
            continue
        details.append(item)
        key = item["creative_key"]
        if key not in records:
            records[key] = dict(item)
            records[key]["ads_count"] = 1
            records[key]["account_ids"] = {item["account_id"]}
            records[key]["account_names"] = {item["account_name"]}
            records[key]["ad_names"] = {item["ad_name"]}
            records[key]["campaign_names"] = {item["campaign_name"]}
            records[key]["adset_names"] = {item["adset_name"]}
            records[key]["schedule_reasons"] = {item["schedule_reason"]}
        else:
            merge_creative_record(records[key], item)
    return records, details


def aggregate_details_by_creative(details: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for item in details:
        key = item["creative_key"]
        if key not in records:
            records[key] = dict(item)
            records[key]["ads_count"] = 1
            records[key]["account_ids"] = {item["account_id"]}
            records[key]["account_names"] = {item["account_name"]}
            records[key]["ad_names"] = {item["ad_name"]}
            records[key]["campaign_names"] = {item["campaign_name"]}
            records[key]["adset_names"] = {item["adset_name"]}
            records[key]["schedule_reasons"] = {item["schedule_reason"]}
        else:
            merge_creative_record(records[key], item)
    return records


def upsert_creative(conn: sqlite3.Connection, record: dict[str, Any], now: str) -> None:
    existing = conn.execute(
        "SELECT creative_key FROM creatives WHERE creative_key = ?",
        (record["creative_key"],),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE creatives
            SET last_seen_at = ?,
                creative_name = COALESCE(NULLIF(?, ''), creative_name),
                material_type = COALESCE(NULLIF(?, ''), material_type),
                material_label = COALESCE(NULLIF(?, ''), material_label),
                thumbnail_url = COALESCE(NULLIF(?, ''), thumbnail_url),
                image_hash = COALESCE(NULLIF(?, ''), image_hash),
                video_id = COALESCE(NULLIF(?, ''), video_id),
                raw_json = ?
            WHERE creative_key = ?
            """,
            (
                now,
                record["creative_name"],
                record["material_type"],
                record["material_label"],
                record["thumbnail_url"],
                record["image_hash"],
                record["video_id"],
                record["raw_json"],
                record["creative_key"],
            ),
        )
        return

    conn.execute(
        """
        INSERT INTO creatives (
            creative_key, creative_id, creative_name, material_type, material_label,
            thumbnail_url, image_hash, video_id, object_type, first_seen_at,
            first_seen_date, last_seen_at, first_account_id, first_ad_id, first_ad_name,
            first_campaign_name, first_adset_name, meta_created_time, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            record["creative_key"],
            record["creative_id"],
            record["creative_name"],
            record["material_type"],
            record["material_label"],
            record["thumbnail_url"],
            record["image_hash"],
            record["video_id"],
            record["object_type"],
            now,
            now[:10],
            now,
            record["account_id"],
            record["ad_id"],
            record["ad_name"],
            record["campaign_name"],
            record["adset_name"],
            record["created_time"],
            record["raw_json"],
        ),
    )


def upsert_metric_rows(
    conn: sqlite3.Connection,
    account_id: str,
    creative_by_ad: dict[str, str],
    daily_rows: list[dict[str, Any]],
    now: str,
) -> None:
    for row in daily_rows:
        ad_id = str(row.get("ad_id") or "")
        creative_key = creative_by_ad.get(ad_id)
        if not ad_id or not creative_key:
            continue
        conn.execute(
            """
            INSERT INTO ad_daily_metrics (
                account_id, ad_id, creative_key, date_start, date_stop,
                ad_name, adset_id, adset_name, campaign_id, campaign_name,
                spend, impressions, clicks, leads, actions_json, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(account_id, ad_id, creative_key, date_start, date_stop)
            DO UPDATE SET
                ad_name = excluded.ad_name,
                adset_id = excluded.adset_id,
                adset_name = excluded.adset_name,
                campaign_id = excluded.campaign_id,
                campaign_name = excluded.campaign_name,
                spend = excluded.spend,
                impressions = excluded.impressions,
                clicks = excluded.clicks,
                leads = excluded.leads,
                actions_json = excluded.actions_json,
                updated_at = excluded.updated_at
            """,
            (
                account_id,
                ad_id,
                creative_key,
                row["date_start"],
                row["date_stop"],
                row["ad_name"],
                row["adset_id"],
                row["adset_name"],
                row["campaign_id"],
                row["campaign_name"],
                row["spend"],
                row["impressions"],
                row["clicks"],
                row["leads"],
                row["actions_json"],
                now,
            ),
        )


def upsert_seen_ad(conn: sqlite3.Connection, detail: dict[str, Any], now: str) -> None:
    conn.execute(
        """
        INSERT INTO ads_seen (
            ad_id, account_id, creative_key, first_seen_at, last_seen_at,
            ad_name, campaign_name, adset_name, created_time, updated_time,
            adset_start_time, campaign_start_time, schedule_reason,
            effective_status, raw_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(ad_id)
        DO UPDATE SET
            last_seen_at = excluded.last_seen_at,
            creative_key = excluded.creative_key,
            ad_name = excluded.ad_name,
            campaign_name = excluded.campaign_name,
            adset_name = excluded.adset_name,
            updated_time = excluded.updated_time,
            adset_start_time = excluded.adset_start_time,
            campaign_start_time = excluded.campaign_start_time,
            schedule_reason = excluded.schedule_reason,
            effective_status = excluded.effective_status,
            raw_json = excluded.raw_json
        """,
        (
            detail["ad_id"],
            detail["account_id"],
            detail["creative_key"],
            now,
            now,
            detail["ad_name"],
            detail["campaign_name"],
            detail["adset_name"],
            detail["created_time"],
            detail["updated_time"],
            detail["adset_start_time"],
            detail["campaign_start_time"],
            detail["schedule_reason"],
            detail["effective_status"],
            detail["raw_json"],
        ),
    )


def record_run(
    conn: sqlite3.Connection,
    started_at: str,
    since: date,
    until: date,
    accounts: list[str],
    baseline_only: bool,
    discovered: int,
    reported: int,
    report_dir: Path,
) -> None:
    conn.execute(
        """
        INSERT INTO runs (
            started_at, since_date, until_date, accounts_json, baseline_only,
            discovered_creatives, reported_new_creatives, report_dir
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            started_at,
            since.isoformat(),
            until.isoformat(),
            json.dumps(accounts, ensure_ascii=False),
            1 if baseline_only else 0,
            discovered,
            reported,
            str(report_dir),
        ),
    )


SUMMARY_COLUMNS = [
    "creative_key",
    "creative_id",
    "creative_name",
    "region",
    "material_label",
    "material_type",
    "account_ids",
    "account_names",
    "ads_count",
    "spend",
    "impressions",
    "clicks",
    "ctr",
    "leads",
    "campaign_names",
    "adset_names",
    "ad_names",
    "created_time",
    "schedule_reasons",
    "thumbnail_url",
]

DETAIL_COLUMNS = [
    "creative_key",
    "account_id",
    "account_name",
    "region",
    "material_label",
    "material_type",
    "creative_id",
    "creative_name",
    "ad_id",
    "ad_name",
    "campaign_id",
    "campaign_name",
    "adset_id",
    "adset_name",
    "created_time",
    "updated_time",
    "adset_start_time",
    "campaign_start_time",
    "schedule_reason",
    "effective_status",
    "spend",
    "impressions",
    "clicks",
    "ctr",
    "leads",
    "thumbnail_url",
]


def compact_record(record: dict[str, Any]) -> dict[str, Any]:
    out = dict(record)
    for key in ("account_ids", "account_names", "ad_names", "campaign_names", "adset_names", "schedule_reasons"):
        value = out.get(key)
        if isinstance(value, set):
            out[key] = " | ".join(sorted(v for v in value if v))
    out["spend"] = round(as_float(out.get("spend")), 2)
    out["ctr"] = round(as_float(out.get("ctr")) * 100, 2)
    out["leads"] = round(as_float(out.get("leads")), 2)
    return out


def compact_detail(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    out["spend"] = round(as_float(out.get("spend")), 2)
    out["ctr"] = round(as_float(out.get("ctr")) * 100, 2)
    out["leads"] = round(as_float(out.get("leads")), 2)
    return out


def write_outputs(
    output_base: Path,
    since: date,
    until: date,
    accounts: list[str],
    new_records: list[dict[str, Any]],
    new_details: list[dict[str, Any]],
    errors: list[str],
    baseline_only: bool,
    min_spend_alert: float,
) -> Path:
    report_dir = output_base / datetime.now().strftime("%Y%m%d_%H%M%S")
    report_dir.mkdir(parents=True, exist_ok=True)
    summary_rows = [compact_record(r) for r in sorted(new_records, key=lambda x: (-x["spend"], x["region"], x["creative_name"]))]
    detail_rows = [compact_detail(r) for r in sorted(new_details, key=lambda x: (x["creative_key"], x["account_id"], x["ad_id"]))]

    write_csv(report_dir / "new_creatives.csv", SUMMARY_COLUMNS, summary_rows)
    write_csv(report_dir / "new_creative_ad_details.csv", DETAIL_COLUMNS, detail_rows)
    markdown = build_markdown(since, until, accounts, summary_rows, errors, baseline_only, min_spend_alert)
    (report_dir / "fb_new_creatives_broadcast.md").write_text(markdown, encoding="utf-8")
    if Workbook:
        write_xlsx(report_dir / "fb_new_creatives_broadcast.xlsx", markdown, summary_rows, detail_rows)
    return report_dir


def write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_xlsx(path: Path, markdown: str, summary_rows: list[dict[str, Any]], detail_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "播报"
    for idx, line in enumerate(markdown.splitlines(), start=1):
        ws.cell(row=idx, column=1, value=line)
    write_sheet(wb, "新素材", SUMMARY_COLUMNS, summary_rows)
    write_sheet(wb, "广告明细", DETAIL_COLUMNS, detail_rows)
    wb.save(path)


def write_sheet(wb: Any, title: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)


def build_markdown(
    since: date,
    until: date,
    accounts: list[str],
    rows: list[dict[str, Any]],
    errors: list[str],
    baseline_only: bool,
    min_spend_alert: float,
) -> str:
    title = "FB广告上新/排期播报"
    if baseline_only:
        title = "FB广告上新/排期监控基线"
    lines = [f"# {title}｜{since.isoformat()} 至 {until.isoformat()}", ""]
    lines.append(f"- 监控账户：{len(accounts)} 个")
    lines.append("- 主口径：14天内新建的广告，或广告组排期在窗口内/未来排期的广告；用 ad_id 防重复播报。")
    lines.append("- 素材汇总口径：同一批广告再按 creative_id / video_id / image_hash / 素材名聚合。")
    lines.append("- 指标口径：Meta Insights API，线索/约课按 action_type 映射统计。")
    if baseline_only:
        lines.append("- 本次是 baseline-only：只建立历史库，不作为正式新增播报。")
    lines.append("")

    if not rows:
        lines.append("## 概览")
        lines.append("")
        lines.append("- 本次没有发现需要播报的新广告/排期广告。")
        append_errors(lines, errors)
        return "\n".join(lines) + "\n"

    by_region = Counter(row["region"] for row in rows)
    by_type = Counter(row["material_label"] for row in rows)
    total_ads = sum(as_int(row.get("ads_count")) for row in rows)
    started = [row for row in rows if as_float(row["spend"]) > 0 or as_int(row["impressions"]) > 0]
    zero = len(rows) - len(started)
    high_spend_no_lead = [
        row for row in rows if as_float(row["spend"]) >= min_spend_alert and as_float(row["leads"]) <= 0
    ]
    naming_issues = [row for row in rows if row["region"] == "未识别" or row["material_label"] == "未识别"]

    lines.append("## 概览")
    lines.append("")
    lines.append(f"- 上新/排期广告：{total_ads} 条")
    lines.append(f"- 涉及素材：{len(rows)} 个")
    lines.append(f"- 已起量素材：{len(started)} 个；无消耗/无展示素材：{zero} 个")
    lines.append(f"- 总花费：{sum(as_float(r['spend']) for r in rows):.2f}")
    lines.append(f"- 总展示：{sum(as_int(r['impressions']) for r in rows)}")
    lines.append(f"- 总点击：{sum(as_int(r['clicks']) for r in rows)}")
    lines.append(f"- 线索/约课：{sum(as_float(r['leads']) for r in rows):.2f}")
    lines.append("")

    lines.append("## 分布")
    lines.append("")
    lines.append("- 地区：" + "；".join(f"{k} {v}" for k, v in by_region.most_common()))
    lines.append("- 类型：" + "；".join(f"{k} {v}" for k, v in by_type.most_common()))
    lines.append("")

    lines.append("## 重点提醒")
    lines.append("")
    if high_spend_no_lead:
        lines.append(f"- 花费超过 {min_spend_alert:.0f} 且暂无线索/约课：{len(high_spend_no_lead)} 个")
    else:
        lines.append(f"- 暂无花费超过 {min_spend_alert:.0f} 且 0 线索/约课的新素材。")
    if naming_issues:
        lines.append(f"- 命名/分类待确认：{len(naming_issues)} 个，主要是地区或素材类型未识别。")
    else:
        lines.append("- 地区和素材类型均已识别。")
    lines.append("")

    lines.append("## 上新/排期素材清单")
    lines.append("")
    lines.append("|地区|类型|纳入原因|素材名|账户|广告数|花费|展示|点击|CTR%|线索/约课|缩略图|")
    lines.append("|---|---|---|---|---:|---:|---:|---:|---:|---:|---:|---|")
    for row in rows[:80]:
        name = escape_md(shorten(row["creative_name"] or row["ad_names"] or row["creative_key"], 42))
        thumb = row.get("thumbnail_url") or ""
        thumb_cell = f"[查看]({thumb})" if thumb else ""
        lines.append(
            "|{region}|{label}|{reason}|{name}|{accounts}|{ads}|{spend:.2f}|{impressions}|{clicks}|{ctr:.2f}|{leads:.2f}|{thumb}|".format(
                region=escape_md(row["region"]),
                label=escape_md(row["material_label"]),
                reason=escape_md(shorten(row.get("schedule_reasons", ""), 24)),
                name=name,
                accounts=escape_md(shorten(row["account_names"], 24)),
                ads=as_int(row["ads_count"]),
                spend=as_float(row["spend"]),
                impressions=as_int(row["impressions"]),
                clicks=as_int(row["clicks"]),
                ctr=as_float(row["ctr"]),
                leads=as_float(row["leads"]),
                thumb=thumb_cell,
            )
        )
    if len(rows) > 80:
        lines.append(f"\n> 清单超过 80 行，完整结果见 CSV/XLSX。")

    append_errors(lines, errors)
    return "\n".join(lines) + "\n"


def append_errors(lines: list[str], errors: list[str]) -> None:
    if not errors:
        return
    lines.append("")
    lines.append("## 拉取异常")
    lines.append("")
    for err in errors:
        lines.append(f"- {err}")


def shorten(value: str, length: int) -> str:
    text = str(value or "")
    if len(text) <= length:
        return text
    return text[: length - 1] + "..."


def escape_md(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ")


def main() -> int:
    args = parse_args()
    config = load_config(args)
    since, until = resolve_window(args)
    started_at = datetime.now().isoformat(timespec="seconds")
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    client = FBClient(config.access_token, config.api_version)

    if args.check_only:
        for account_id in config.accounts:
            name = fetch_account_name(client, account_id)
            print(f"[OK] {account_id} {name}")
        return 0

    db_path = Path(args.db)
    if not db_path.is_absolute():
        db_path = SCRIPT_DIR / db_path
    output_base = Path(args.output_dir)
    if not output_base.is_absolute():
        output_base = SCRIPT_DIR / output_base

    conn = None if args.no_save else connect_db(db_path)
    existing_creatives = existing_creative_keys(conn)
    existing_ads = existing_ad_ids(conn)
    lead_types = action_types(args)
    all_records: dict[str, dict[str, Any]] = {}
    all_details: list[dict[str, Any]] = []
    all_daily_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    creative_by_account_ad: dict[tuple[str, str], str] = {}
    errors: list[str] = []

    for account_id in config.accounts:
        try:
            account_name = fetch_account_name(client, account_id)
            print(f"[INFO] account {account_id}: {account_name}", flush=True)

            if args.source_mode == "activity":
                insight_rows = fetch_insights(
                    client,
                    account_id,
                    since,
                    until,
                    daily=args.daily_insights,
                    verbose=args.verbose,
                )
                metrics_by_ad, daily_rows = summarize_insights(insight_rows, lead_types)
                ads, ad_fetch_errors = fetch_ads_by_ids(client, sorted(metrics_by_ad))
                errors.extend(f"{account_id} {err}" for err in ad_fetch_errors)

                if not args.skip_zero_spend_scan:
                    recent_ads, truncated = fetch_recent_ads(
                        client,
                        account_id,
                        since,
                        until,
                        args.max_ads_per_account,
                        verbose=args.verbose,
                    )
                    if truncated:
                        errors.append(f"{account_id}: zero-spend ad scan hit max-ads-per-account={args.max_ads_per_account}")
                    for ad in recent_ads:
                        ad_id = str(ad.get("id") or "")
                        if ad_id and ad_id not in ads:
                            ads[ad_id] = ad
                source_count = len(metrics_by_ad)
                source_label = "insights_ads"
            else:
                candidate_ads, truncated, scan_errors = fetch_created_or_scheduled_ads(
                    client,
                    account_id,
                    since,
                    until,
                    args.future_scheduled_days,
                    args.max_ads_per_account,
                    verbose=args.verbose,
                )
                errors.extend(f"{account_id} {err}" for err in scan_errors)
                if truncated:
                    errors.append(f"{account_id}: ad scan hit max-ads-per-account={args.max_ads_per_account}")
                ads = {str(ad.get("id")): ad for ad in candidate_ads if ad.get("id")}
                if args.skip_insights:
                    metrics_by_ad = {}
                    daily_rows = []
                else:
                    insight_rows = fetch_insights(
                        client,
                        account_id,
                        since,
                        until,
                        daily=args.daily_insights,
                        verbose=args.verbose,
                    )
                    metrics_by_ad, daily_rows = summarize_insights(insight_rows, lead_types)
                source_count = len(ads)
                source_label = "created_or_scheduled_ads"

            account_records, account_details = build_records(
                account_id,
                account_name,
                ads,
                metrics_by_ad,
                include_metric_only=(args.source_mode == "activity"),
            )
            for key, record in account_records.items():
                if key not in all_records:
                    all_records[key] = record
                else:
                    merge_creative_record(all_records[key], record)
            all_details.extend(account_details)
            all_daily_rows[account_id].extend(daily_rows)
            for detail in account_details:
                creative_by_account_ad[(account_id, detail["ad_id"])] = detail["creative_key"]
            print(
                f"[INFO] {account_id}: {source_label}={source_count} discovered_ads={len(account_details)} discovered_creatives={len(account_records)}",
                flush=True,
            )
        except GraphAPIError as exc:
            errors.append(f"{account_id}: {exc}")

    if conn is not None:
        for record in all_records.values():
            upsert_creative(conn, record, now)
        if args.source_mode == "created-or-scheduled":
            for detail in all_details:
                upsert_seen_ad(conn, detail, now)
        for account_id, rows in all_daily_rows.items():
            ad_map = {
                ad_id: key
                for (acct, ad_id), key in creative_by_account_ad.items()
                if acct == account_id
            }
            upsert_metric_rows(conn, account_id, ad_map, rows, now)

    if args.baseline_only:
        report_records: list[dict[str, Any]] = []
        report_details: list[dict[str, Any]] = []
    elif args.source_mode == "created-or-scheduled":
        new_ad_ids = {row["ad_id"] for row in all_details} - existing_ads
        report_details = [row for row in all_details if row["ad_id"] in new_ad_ids]
        report_records = list(aggregate_details_by_creative(report_details).values())
    else:
        new_keys = set(all_records) - existing_creatives
        report_records = [all_records[key] for key in new_keys]
        report_details = [row for row in all_details if row["creative_key"] in new_keys]

    report_dir = write_outputs(
        output_base,
        since,
        until,
        config.accounts,
        report_records,
        report_details,
        errors,
        args.baseline_only,
        args.min_spend_alert,
    )

    if conn is not None:
        record_run(
            conn,
            started_at,
            since,
            until,
            config.accounts,
            args.baseline_only,
            len(all_records),
            len(report_records),
            report_dir,
        )
        conn.commit()
        conn.close()

    print(f"[OK] discovered ads: {len(all_details)}", flush=True)
    print(f"[OK] discovered creatives: {len(all_records)}", flush=True)
    if args.baseline_only:
        print("[OK] baseline saved; no new creative broadcast emitted.", flush=True)
    else:
        if args.source_mode == "created-or-scheduled":
            print(f"[OK] new/scheduled ads reported: {len(report_details)}", flush=True)
            print(f"[OK] involved creatives reported: {len(report_records)}", flush=True)
        else:
            print(f"[OK] new creatives reported: {len(report_records)}", flush=True)
    print(f"[OK] report dir: {report_dir}", flush=True)
    if errors:
        print(f"[WARN] completed with {len(errors)} pull warnings. See Markdown report.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
