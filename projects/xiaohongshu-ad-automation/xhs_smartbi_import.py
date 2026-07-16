#!/usr/bin/env python3
"""Normalize the SmartBI Xiaohongshu material-format workbook for xhs_loop."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT = ROOT / "inputs" / "performance_data" / "smartbi_xhs_format_summary.csv"


def number(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def merged_headers(sheet) -> list[str]:
    top = [cell.value for cell in sheet[5]]
    bottom = [cell.value for cell in sheet[6]]
    return [str(secondary or primary or f"column_{index}").replace("\n", " ").strip() for index, (primary, secondary) in enumerate(zip(top, bottom), start=1)]


def normalize(source: Path, output: Path) -> int:
    sheet = load_workbook(source, data_only=True).active
    headers = merged_headers(sheet)
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=7, values_only=True):
        record = dict(zip(headers, values))
        material_type = str(record.get("素材类型") or "").strip()
        if material_type in {"", "总计"}:
            continue
        rows.append({
            "content_type": material_type,
            "topic": "SmartBI素材类型汇总",
            "title": material_type,
            "views": number(record.get("曝光数")),
            "leads": number(record.get("约课数")),
            "cost": number(record.get("消耗")),
            "cpc": number(record.get("CPC")),
            "ctr": number(record.get("CTR")),
            "cvr": number(record.get("CVR")),
            "lead_cost": number(record.get("约课成本")),
            "roi2": number(record.get("当月ROI2")),
            "source_period": str(record.get("月份") or ""),
            "source_level": "素材类型汇总",
        })
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = list(rows[0]) if rows else ["content_type", "source_level"]
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    print(f"normalized_rows={normalize(args.source, args.output)}")
    print(f"output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
